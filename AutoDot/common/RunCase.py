# coding:utf-8
__author__ = 'GRUNMI'


from common.GetProjectCase import GetCase
from common.LogOutput import Log
from common.HttpRequest import HttpRequest
import openpyxl
import time
import os
from common.DataResultCheck import data_result_check
from common.CreatePieChart import CreatePie
from common.GetRunProjectCaseNumResult import get_project_run_num
from common.CreateHtmlReport import create_html_report
from common.ProjectReport import get_html_project_report
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def run_case(projectname, type=0, id=None):
    '''
    :type 0代表项目用例，1代表场景用例
    :id 用例id
    :生成用例规则：场景用例需要在报告前加上 settingReport
    '''
    mylogger = Log().get_logger()
    mylogger.info('当前运行文件：{}'.format(__file__))
    mylogger.info('获取{}项目的所有用例'.format(projectname))
    project = GetCase(projectname).get_project_case()
    all_case = project[0]
    case_file = project[2]
    case_name = project[3]

    runResult = []

    try:
        if all_case:
            wb = openpyxl.load_workbook(case_file)
            ws = wb.active
            startTime = time.time()
            if type == 0:
                if str(all_case).lower().find('yes') < 0:
                    return 'noRunCase'
                for case in all_case:
                    if case['Run'].lower() == 'yes':
                        mylogger.info('{}项目的第{}行用例，是否执行为：{}'.format(projectname, case['Sort'], case['Run']))
                        mylogger.info('准备运行用例')
                        url = case['Protocol']+case['Host']+case['Path']
                        method = case['Method']
                        headers = case['Headers']
                        try:
                            result = HttpRequest().send_request(url=url, method=method, headers=headers)
                            # 写入到Response单元格
                            ws['R' + str(case['RowID'])] = result.text
                            # 获取Response值
                            Response = result.text
                        except:
                            Response = ws['R' + str(case['RowID'])] = ''
                        # 校验值
                        # 获取Expect值
                        Except = case['Expect']
                        # 获取CheckPoint值
                        CheckPoint = case['CheckPoint']

                        # 对比数据
                        checkResult = data_result_check(Except, CheckPoint, Response)

                        if checkResult[0] == checkResult[1] and Response:
                            testResult = ws['Z' + str(case['RowID'])] = 'true'

                        else:
                            testResult = ws['Z' + str(case['RowID'])] = 'false'

                        reportResult = {
                            "moduleName": case['CaseName'],
                            "path": case['Path'],
                            "result": testResult,
                            "checkDetail": [
                                {
                                    "预期结果": checkResult[0]
                                },
                                {
                                    "运行结果": checkResult[1]
                                }
                            ]
                        }
                        runResult.append(reportResult)
                    else:
                        mylogger.info('{}项目的第{}行用例，是否执行为：{}，不执行用例'.format(projectname, case['RowID'], case['Run']))
            elif type == 1:
                select_id = []
                for i in id.split(','):
                    if i:
                        select_id.append(int(i))
                all_id = []
                for i in all_case:
                    all_id.append(int(i["ID"]))

                # 检测选择的用例是否包含在用例中，在选择过程中，用例有可能被删除
                if set(select_id) <= set(all_id):
                    # 转换成list，去掉最后一个逗号
                    for id in select_id:
                        for case in all_case:
                            # excel取出来是一个float类型
                            if int(id) == int(case['ID']):
                                mylogger.info('{}项目的第{}行用例，是否执行为：{}'.format(projectname, case['Sort'], case['Run']))
                                mylogger.info('准备运行用例')
                                url = case['Protocol'] + case['Host'] + case['Path']
                                method = case['Method']
                                headers = case['Headers']
                                result = HttpRequest().send_request(url=url, method=method, headers=headers)
                                # 写入到Response单元格
                                ws['R' + str(case['RowID'])] = result.text
                                # 校验值
                                # 获取Expect值
                                Except = case['Expect']
                                # 获取CheckPoint值
                                CheckPoint = case['CheckPoint']
                                # 获取Response值
                                # Response = ws['R' + str(case['RowID'])]
                                Response = result.text
                                # 对比数据
                                checkResult = data_result_check(Except, CheckPoint, Response)
                                if checkResult[0] == checkResult[1]:
                                    testResult = ws['Z' + str(case['RowID'])] = 'true'
                                else:
                                    testResult = ws['Z' + str(case['RowID'])] = 'false'

                                reportResult = {
                                    "moduleName": case['CaseName'],
                                    "path": case['Path'],
                                    "result": testResult,
                                    "checkDetail": [
                                        {
                                            "预期结果": checkResult[0]
                                        },
                                        {
                                            "运行结果": checkResult[1]
                                        }
                                    ]
                                }
                                runResult.append(reportResult)
                else:
                    difference_value = list(set(select_id).difference(set(all_id)))
                    mylogger.info("勾选的ID{}，用例表中的ID{}，表中不存在勾选的ID：{}".format(select_id, all_id, difference_value))
                    return 'selectFalse'

            endTime = time.time()
            reportFolder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'report\\', case_name, 'excelReport')
            if not os.path.exists(reportFolder):
                os.makedirs(reportFolder)
                mylogger.info("生成用例目录路径：{}".format(reportFolder))

            if type == 0:
                reportName = time.strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
            elif type == 1:
                reportName = 'settingReport' + time.strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
            reportCaseFile = os.path.join(reportFolder, reportName)
            wb.save(reportCaseFile)

            # 在excel中生成图表报告
            if type == 0:
                CreatePie(projectname, type=0).pie()
            elif type == 1:
                CreatePie(projectname, type=1).pie()

            # 生成html图形报告
            if type == 0:
                runNum = get_project_run_num(projectname, type=0)
                success = runNum[0]
                failed = runNum[1]
                sum = success+failed

                str(success / sum).count('.') == 1
                right = str(success / sum).split('.')[1]
                if len(right) > 2:
                    passRate = str(round(success / sum * 100, 1)) + '%'
                    failedRate = str(round((1 - success / sum) * 100, 1)) + '%'
                else:
                    passRate = str(int(success / sum * 100)) + '%'
                    failedRate = str(int((1 - success / sum) * 100)) + '%'

                from platform import python_version
                pythonVersion = python_version()
                import django
                djangoVersion = django.get_version()
                data = {
                    "projectName": projectname,
                    "testPersonName": "GRUNMI",
                    "startTime": time.strftime('%Y-%m-%d %H:%M:%S'),
                    "runTime": endTime-startTime,
                    "pass": str(success),
                    "failed": str(failed),
                    "sum": sum,
                    "passRate": passRate,
                    "failedRate": failedRate,
                    "runEnvironment": "python%s+django%s" % (pythonVersion, djangoVersion),
                    "runResult": runResult
                }
                create_html_report(data, type=0)
                # get_html_project_report(projectname, type=0)
            elif type == 1:
                runNum = get_project_run_num(projectname, type=1)
                success = runNum[0]
                failed = runNum[1]
                sum = runNum[0] + failed

                str(success / sum).count('.') == 1
                right = str(success / sum).split('.')[1]
                if len(right) > 2:
                    passRate = str(round(success / sum * 100, 1)) + '%'
                    failedRate = str(round((1 - success / sum) * 100, 1)) + '%'
                else:
                    passRate = str(int(success / sum * 100)) + '%'
                    failedRate = str(int((1 - success / sum) * 100)) + '%'

                from platform import python_version
                pythonVersion = python_version()
                import django
                djangoVersion = django.get_version()
                data = {
                    "projectName": projectname,
                    "testPersonName": "GRUNMI",
                    "startTime": time.strftime('%Y-%m-%d %H:%M:%S'),
                    "runTime": endTime - startTime,
                    "pass": str(success),
                    "failed": str(failed),
                    "sum": sum,
                    "passRate": passRate,
                    "failedRate": failedRate,
                    "runEnvironment": "python%s+django%s" % (pythonVersion, djangoVersion),
                    "runResult": runResult
                }
                create_html_report(data, type=1)
                # get_html_project_report(projectname, type=1)
            return 'success'
        else:
            return 'caseEmpty'
    except Exception as e:
        mylogger.error('运行错误，请查看错误信息：'.format(e))
        return 'unknownError', e

if __name__ == '__main__':
    test = run_case('1212')
    print(test)


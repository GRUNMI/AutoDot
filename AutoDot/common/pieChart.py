# coding:utf-8

__author__ = 'GRUNMI'


import openpyxl
from openpyxl.chart import PieChart, Reference
import xlrd
import os
import time
import common.logOutput



# 获取文件中的运行结果数据
def get_test_report_num():
    PATH = lambda P: os.path.abspath(os.path.join(os.path.dirname(__file__), P))
    caseFolder = PATH("../report")
    for root, dirs, files in os.walk(caseFolder):
        # print(files[-1::])
        for name in files[-1::]:
            caseFile = os.path.join(caseFolder, name)
            wb = xlrd.open_workbook(caseFile)
            ws = wb.sheet_by_index(0)
            contents = ws.col_values(23)
            print(contents)
            print(contents.count(''))
            noRunNum = (contents.count(''))
            print(contents.count('true'))
            successNum = (contents.count('true'))
            print(contents.count('false'))
            failedNum = (contents.count('false'))
            print(successNum,failedNum,noRunNum, caseFile)

# 生成图表
def pie(successNum, failedNum, noRunNum, reportCaseFile):
    mylogger = common.logOutput.Log().get_logger()
    mylogger.info("------------------------------------------------------------------")
    mylogger.info("当前运行文件：{}".format(__file__))
    try:
        mylogger.info("准备打开用例文件，准备生成图表")
        wb = openpyxl.load_workbook(reportCaseFile)
        mylogger.info("成功用例数：{}个，失败用例数：{}个，不执行用例数：{}个 ".format(successNum, failedNum, noRunNum))
        rows = [
            ['结果', '数量'],
            ['成功', successNum],
            ['失败', failedNum],
            ['不执行', noRunNum],
        ]
        ws = wb.create_sheet('casePie')
        # 把值添加在excel中
        for row in rows:
            ws.append(row)

        # 创建图表实例
        chart = PieChart()

        # 设置类别的取值，min_row和max_row有bug，需要加1
        labels = Reference(ws, min_col=1, min_row=2, max_row=5)
        # 设置数据的取值，min_col在哪一行取值
        data = Reference(ws, min_col=2, min_row=1, max_row=4)
        # titles_from_data：鼠标移动上去图表上显示数据
        chart.add_data(data, titles_from_data=True)

        # 设置标题
        chart.title = "用例运行结果图表"
        # 设置类别
        chart.set_categories(labels=labels)
        # 将图表添加在哪个位置
        ws.add_chart(chart, 'D1')
        wb.save(reportCaseFile)
        mylogger.info("生成图表完成，保存在{}文件中".format(reportCaseFile))
    except Exception as e:
        mylogger.debug("出现异常，异常日志：{}".format(e))


if __name__ == '__main__':
    # reportFolder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'report\\')
    # reportName = time.strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
    # if not os.path.exists(reportFolder):
    #     os.mkdir(reportFolder)
    #     print("生成用例目录路径：{}".format(reportFolder))
    # reportCaseFile = os.path.join(reportFolder, reportName)
    # wb = openpyxl.Workbook(reportCaseFile)
    # wb.save(reportCaseFile)
    # print("生成用例文件路径：{}".format(reportCaseFile))

    PATH = lambda P: os.path.abspath(os.path.join(os.path.dirname(__file__), P))

    caseFolder = PATH("../report")
    for root, dirs, files in os.walk(caseFolder):
        # print(files[-1::])
        for name in files[-1::]:
            caseFile = os.path.join(caseFolder, name)
            wb = xlrd.open_workbook(caseFile)
            ws = wb.sheet_by_index(0)
            contents = ws.col_values(23)
            print(contents)
            print(contents.count(''))
            noRunNum = (contents.count(''))
            print(contents.count('true'))
            successNum = (contents.count('true'))
            print(contents.count('false'))
            failedNum = (contents.count('false'))


            pie(successNum, failedNum, noRunNum, reportCaseFile=caseFile)

# from openpyxl.chart import (
#     PieChart,
#     ProjectedPieChart,
#     Reference
# )
# from openpyxl.chart.series import DataPoint
#
# data = [
#     ['Pie', 'Sold'],
#     ['Apple', 50],
#     ['Cherry', 30],
#     ['Pumpkin', 10],
#     ['Chocolate', 40],
# ]
#
# wb = openpyxl.load_workbook(case_file)
# ws = wb.active
# print(ws)
# for row in data:
#     ws.append(row)
#
# pie = PieChart()
# labels = Reference(ws, min_col=1, min_row=1, max_row=5)
# data = Reference(ws, min_col=2, min_row=1, max_row=5)
# pie.add_data(data, titles_from_data=True)
# pie.set_categories(labels)
# pie.title = "Pies sold by category"
#
# # Cut the first slice out of the pie
#
# # slice = DataPoint(idx=0, explosion=20)
# # pie.series[0].data_points = [slice]
#
# ws.add_chart(pie, "D1")
#
#
# # ws = wb.create_sheet(title="Projection")
# #
# # data = [
# #     ['Page', 'Views'],
# #     ['Search', 95],
# #     ['Products', 4],
# #     ['Offers', 0.5],
# #     ['Sales', 0.5],
# # ]
# #
# # for row in data:
# #     ws.append(row)
# #
# # projected_pie = ProjectedPieChart()
# # projected_pie.type = "pie"
# # projected_pie.splitType = "val" # split by value
# # labels = Reference(ws, min_col=1, min_row=2, max_row=5)
# # data = Reference(ws, min_col=2, min_row=1, max_row=5)
# # projected_pie.add_data(data, titles_from_data=True)
# # projected_pie.set_categories(labels)
# #
# # ws.add_chart(projected_pie, "A10")
# #
# # from copy import deepcopy
# # projected_bar = deepcopy(projected_pie)
# # projected_bar.type = "bar"
# # projected_bar.splitType = 'pos' # split by position
# #
# # ws.add_chart(projected_bar, "A27")
#
# wb.save(case_file)




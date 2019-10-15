# coding:utf-8
__author__ = 'GRUNMI'


import os
import openpyxl
from openpyxl.styles import Font
from common.logOutput import Log


def isExistCase(projectName):
    mylogger = Log().get_logger()
    mylogger.info("------------------------------------------------------------------")
    mylogger.info("当前运行文件：{}".format(__file__))
    caseFolder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))+'\\case\\')
    # print(caseFolder)
    caseFile = os.path.join(caseFolder, projectName + '.xlsx')

    # 判断是否存在case目录，没有，则创建
    if not os.path.exists(caseFolder):
        os.mkdir(caseFolder)
        mylogger.info("生成用例目录路径：{}".format(caseFolder))

    # 判断是否存在case模板，没有，则创建
    if not os.path.exists(caseFile):
        wb = openpyxl.Workbook(caseFile)
        wb.create_sheet(title=projectName+'接口用例')
        mylogger.info("生成用例文件路径：{}".format(caseFile))
        wb.save(caseFile)
        wb = openpyxl.load_workbook(caseFile, data_only=True)
        ws = wb.active
        mylogger.info("创建用例标题")
        ws['A1'] = 'ID'
        ws['B1'] = 'Run'
        ws['C1'] = 'ModuleName'
        ws['D1'] = 'CaseName'
        ws['E1'] = 'CaseDescription'
        ws['F1'] = 'Protocol'
        ws['G1'] = 'Host'
        ws['H1'] = 'Method'
        ws['I1'] = 'Headers'
        ws['J1'] = 'Path'
        ws['K1'] = 'ReplaceID'
        ws['L1'] = 'Data'
        ws['M1'] = 'Files'
        ws['N1'] = 'Expect'
        ws['O1'] = 'CheckPoint'
        ws['P1'] = 'Result'
        ws['Q1'] = 'TransmitID'
        ws['R1'] = 'TransmitTargetID'
        ws['S1'] = 'Remark'
        ws['T1'] = 'CreateDate'
        ws['U1'] = 'UpdateTime'
        ws['V1'] = 'Author'
        ws['W1'] = 'Editor'
        ws['X1'] = 'run_result'
        # print(ws.rows)
        # print(ws[1])
        mylogger.info("修改用例标题样式：宋体加粗")
        for cell in ws[1]:
            cell.font = Font('宋体', bold=True)
        # a1 = ws['A1']
        # a1.font = Font('黑体', bold=True)
        wb.save(caseFile)
        mylogger.info("保存创建的用例文件，路径为：{}".format(caseFile))
        wb.close()
    else:
        mylogger.info('已存在{}用例文件：{}'.format(projectName, caseFile))
    return caseFile

if __name__ == '__main__':
    excel = isExistCase('text12')
    print(excel)
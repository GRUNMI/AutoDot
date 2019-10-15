# coding:utf-8
__author__ = 'GRUNMI'

from common.isExistTestProject import isExistProject
from common.logOutput import Log
import openpyxl
import time


def apiData(projectName,action,rowid=1,Run='YES',ModuleName='模块名称',CaseName='用例名称',CaseDescription='用例描述',protocol='http://',host=None,method='GET',response=None,path=None,headers=None,data=None,file=None,replaceID=None,CheckPoint=None,TransmitID=None,TransmitTargetID=None):
    '''


    :param projectName: 项目名称
    :param action: 执行动作，1增 2删 3修改用例数据 4修改用例执行状态 5查
    :param rowid: 用例所在行数位置,删除时必填
    :param Run: 用例是否执行
    :param ModuleName: 用例模块名称
    :param CaseName: 用例名称
    :param CaseDescription: 用例描述
    :param protocol: 协议
    :param host: 域名
    :param method: 请求方法
    :param response: 响应结果
    :param path: 路径
    :param headers: 数据请求格式
    :param data: 请求数据
    :param file:
    :param replaceID: 替换请求的动态id
    :param CheckPoint:校验数据
    :param TransmitID:保存动态id
    :param TransmitTargetID:保存到动态的目标参数id
    :return:
    '''


    mylogger = Log().get_logger()
    mylogger.info('当前运行文件：{}'.format(__file__))
    mylogger.info('调用运行isExistProject方法')
    caseFile = isExistProject(projectName)[1]
    # print(caseFile)
    mylogger.info('打开{}文件'.format(caseFile))

    action = int(action)
    rowid = int(rowid)
    wb = openpyxl.load_workbook(caseFile, data_only=True)
    # print(wb.get_sheet_names())
    # print(wb.sheetnames)
    # wb.get_sheet_names() 已弃用，需要使用sheetName

    try:
        for ws in wb.sheetnames:
            # count：判断是否包含projectName
            if ws.count(projectName):
                mylogger.info('已匹配到{}的用例表为：{}'.format(projectName, ws))
                # wb.get_sheet_by_name() 已弃用，需要使用wb[]
                ws = wb[ws]
                rows = ws.max_row
                # print(type(rows), '------------------')
                mylogger.info('获取当前{}表的最大行数：{}'.format(ws, rows))
                # 新增
                if action == 1:
                    write_row = str(rows+1)
                    mylogger.info('准备将数据新增在{}表的第{}行'.format(ws, write_row))
                    ws['A' + write_row] = int(write_row) - 1
                    ws['B' + write_row] = Run
                    ws['C' + write_row] = ModuleName
                    ws['D' + write_row] = CaseName
                    ws['E' + write_row] = CaseDescription
                    ws['F' + write_row] = protocol
                    ws['G' + write_row] = host
                    ws['H' + write_row] = method
                    ws['I' + write_row] = headers
                    ws['J' + write_row] = path
                    ws['K' + write_row] = replaceID
                    ws['L' + write_row] = data
                    ws['M' + write_row] = file
                    ws['N' + write_row] = response
                    ws['O' + write_row] = CheckPoint
                    # ws['P' + write_row] = 'Result'
                    ws['Q' + write_row] = TransmitID
                    ws['R' + write_row] = TransmitTargetID
                    ws['S' + write_row] = 'Remark'
                    ws['T' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                    ws['U' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                    ws['V' + write_row] = 'GRUNMI'
                    ws['W' + write_row] = 'GRUNMI'
                    mylogger.info('新增在{}表中第{}行的数据成功'.format(ws, write_row))
                    mylogger.info('总共有{}条用例数据'.format(int(write_row)-1))
                # 删除
                elif action == 2:
                    mylogger.info('准备删除{}表中，第{}行用例数据'.format(ws, rowid))
                    if rowid != 1:
                        try:
                            if not ws['A' + str(rowid)].value:
                                mylogger.info('删除{}表中，第{}行用例数据不存在，不用删除'.format(ws, rowid))
                                break
                            for i in ws['A']:
                                # print(type(ws['A'+str(rowid)].value), ws['A'+str(rowid)].value)
                                # print(type(i.value), i.value)
                                if isinstance(i.value, int):
                                    if i.value > ws['A'+str(rowid)].value:
                                        i.value -= 1
                                        mylogger.info('修改{}表中，第{}行用例的id数据成功'.format(ws, i))
                            ws.delete_rows(rowid)
                            mylogger.info('删除{}表中，第{}行用例数据成功'.format(ws, rowid))
                        except Exception as e:
                            mylogger.info('删除{}表中第{}行用例报错：{}'.format(ws, rowid, e))
                    elif id == 1:
                        mylogger.info('{}表中，删除用例数据不成功，第{}行是用例标题，不可删除'.format(ws, rowid))
                    mylogger.info('总共有{}条用例数据'.format(int(rows) - 1))

                # 修改用例数据
                elif action == 3:
                    mylogger.info('准备修改{}表中，第{}行用例数据'.format(ws, rowid))
                    if rowid != 1:
                        write_row = str(rowid)
                        # ws['A' + write_row] = int(write_row)-1  不需要用例ID修改
                        # ws['B' + write_row] = 'YES'
                        ws['C' + write_row] = 'ModuleName'
                        ws['D' + write_row] = 'CaseName'
                        ws['E' + write_row] = 'CaseDescription'
                        ws['F' + write_row] = protocol
                        ws['G' + write_row] = host
                        ws['H' + write_row] = method
                        ws['I' + write_row] = headers
                        ws['J' + write_row] = path
                        ws['K' + write_row] = replaceID
                        ws['L' + write_row] = data
                        ws['M' + write_row] = file
                        ws['N' + write_row] = response
                        ws['O' + write_row] = CheckPoint
                        # ws['P' + write_row] = 'Result'
                        ws['Q' + write_row] = TransmitID
                        ws['R' + write_row] = TransmitTargetID
                        ws['S' + write_row] = 'Remark'
                        # ws['T' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                        ws['U' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                        # 不需要修改创建人
                        # ws['V' + write_row] = 'GRUNMI'

                        # 需要修改编辑者
                        ws['W' + write_row] = 'GRUNMI'
                        mylogger.info('修改{}表中，第{}行用例数据成功'.format(ws, rowid))
                    elif rowid > rows:
                        mylogger.info('{}表中，修改用例数据不成功，第{}行不存在数据，不用修改'.format(ws, rowid))
                    else:
                        mylogger.info('{}表中，修改用例数据不成功，第{}行是用例标题，不可修改'.format(ws, rowid))

                # 修改是否执行用例
                elif action == 4:
                    mylogger.info('准备修改{}表中，第{}行用例执行状态数据：{}'.format(ws, rowid, Run))
                    if rowid != 1:
                        write_row = str(rowid)
                        ws['B' + write_row] = Run
                        mylogger.info('修改{}表中，第{}行用例执行状态数据更新成功'.format(ws, rowid))
                    elif rowid > rows:
                        mylogger.info('{}表中，修改用例执行状态数据不成功，第{}行不存在数据，不用修改'.format(ws, rowid))
                    else:
                        mylogger.info('{}表中，修改用例数据不成功，第{}行是用例标题，不可修改'.format(ws, rowid))

                # 查询
                elif action == 5:
                    mylogger.info('准备查询{}表中，第{}行用例数据'.format(ws, rowid))
                    if rowid != 1:
                        if not ws['A' + str(rowid)].value:
                            mylogger.info('查询{}表中，第{}行用例数据不存在'.format(ws, rowid))
                            break
                        find_row = str(rowid)

                        ID=ws['A' + find_row].value
                        Run=ws['B' + find_row].value
                        ModuleName=ws['C' + find_row].value
                        CaseName=ws['D' + find_row].value
                        CaseDescription=ws['E' + find_row].value
                        Protocol=ws['F' + find_row].value
                        Host=ws['G' + find_row].value
                        Method=ws['H' + find_row].value
                        Headers=ws['I' + find_row].value
                        Path=ws['J' + find_row].value
                        ReplaceID=ws['K' + find_row].value
                        Data=ws['L' + find_row].value
                        Files=ws['M' + find_row].value
                        Expect=ws['N' + find_row].value
                        CheckPoint=ws['O' + find_row].value
                        Result=ws['P' + find_row].value
                        TransmitID=ws['Q' + find_row].value
                        TransmitTargetID=ws['R' + find_row].value
                        Remark=ws['S' + find_row].value
                        CreateDate=ws['T' + find_row].value
                        UpdateTime=ws['U' + find_row].value
                        Author=ws['V' + find_row].value
                        Editor=ws['W' + find_row].value


                        find_result = {
                            'projectname': projectName,
                            'rowid': rowid,
                            'ID': ID,
                            'Run': Run,
                            'ModuleName': ModuleName,
                            'CaseName': CaseName,
                            'CaseDescription': CaseDescription,
                            'Protocol': Protocol,
                            'Host': Host,
                            'Method': Method,
                            'Headers': Headers,
                            'Path': Path,
                            'ReplaceID': ReplaceID,
                            'Data': Data,
                            'Files': Files,
                            'Expect': Expect,
                            'CheckPoint': CheckPoint,
                            'Result': Result,
                            'TransmitID': TransmitID,
                            'TransmitTargetID': TransmitTargetID,
                            'Remark': Remark,
                            'CreateDate': CreateDate,
                            'UpdateTime': UpdateTime,
                            'Author': Author,
                            'Editor': Editor
                        }
                        mylogger.info('查询{}表中，第{}行用例数据成功：{}'.format(ws, rowid, find_result))
                        wb.close()
                        mylogger.info('关闭{}工作簿成功'.format(projectName))
                        return find_result
                    # elif id > rows:
                    #     mylogger.info('{}表中，查询的第{}行不存在数据'.format(ws, id))
                    else:
                        mylogger.info('{}表中，查询第{}行是用例标题，不是用例数据'.format(ws, rowid))
                mylogger.info('准备保存{}表'.format(ws))
                wb.save(caseFile)
                mylogger.info('保存{}表成功，准备关闭{}工作簿'.format(ws, projectName))
                break

        # mylogger.info('没有匹配到{}的用例表'.format(projectName))
    except Exception as e:
        mylogger.info('不可预知的错误：{}'.format(e))
    wb.close()
    mylogger.info('关闭{}工作簿成功'.format(projectName))

if __name__ == '__main__':
    apiData(projectName='自动干线',action=5,rowid=4,Run='NO',host='hahhaaaaaaaaa',response='hahha',path='hahha',headers='application/x-www-form-urlencoded',data='hahha',file='hahha',replaceID='hahha',CheckPoint='hahha',TransmitID='hahha',TransmitTargetID='hahha')
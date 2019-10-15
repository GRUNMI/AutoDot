# coding:utf-8
__author__ = 'GRUNMI'


import os
import openpyxl
from openpyxl.styles import Font
from common.LogOutput import Log


def create_project(projectName):
    mylogger = Log().get_logger()
    mylogger.info("------------------------------------------------------------------")
    mylogger.info("当前运行文件：{}".format(__file__))
    caseFolder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))+'\\case\\')
    # print(caseFolder)
    caseFile = os.path.join(caseFolder, projectName + '.xlsx')

    # 判断是否存在case目录，没有，则创建
    if not os.path.exists(caseFolder):
        os.mkdir(caseFolder)
        mylogger.info("生成用例目录路径：{}".format(caseFolder))

    # 判断是否存在case模板，没有，则创建
    if not os.path.exists(caseFile):
        wb = openpyxl.Workbook(caseFile)
        wb.create_sheet(title=projectName+'InterfaceData')
        mylogger.info("生成用例文件路径：{}".format(caseFile))
        wb.save(caseFile)
        wb = openpyxl.load_workbook(caseFile, data_only=True)
        ws = wb.active
        mylogger.info("创建用例标题")
        ws['A1'] = 'ID'
        ws['B1'] = 'Run'
        ws['C1'] = 'Sort'
        ws['D1'] = 'Sql'
        ws['E1'] = 'ModuleName'
        ws['F1'] = 'CaseName'
        ws['G1'] = 'CaseDescription'
        ws['H1'] = 'Protocol'
        ws['I1'] = 'Host'
        ws['J1'] = 'Method'
        ws['K1'] = 'Headers'
        ws['L1'] = 'Path'
        ws['M1'] = 'Data'
        ws['N1'] = 'Files'
        ws['O1'] = 'ReplaceID'
        ws['P1'] = 'Expect'
        ws['Q1'] = 'CheckPoint'
        ws['R1'] = 'Response'
        ws['S1'] = 'TransmitID'
        ws['T1'] = 'TransmitTargetID'
        ws['U1'] = 'Remark'
        ws['V1'] = 'CreateDate'
        ws['W1'] = 'UpdateTime'
        ws['X1'] = 'Author'
        ws['Y1'] = 'Editor'
        ws['Z1'] = 'RunResult'
        # print(ws.rows)
        # print(ws[1])
        mylogger.info("修改用例标题样式：宋体加粗")
        for cell in ws[1]:
            cell.font = Font('宋体', bold=True)
        # a1 = ws['A1']
        # a1.font = Font('黑体', bold=True)
        wb.save(caseFile)
        mylogger.info("保存创建的用例文件，路径为：{}".format(caseFile))
        wb.close()
        return 'createSuccess', caseFile
    else:
        mylogger.info('已存在{}用例文件：{}'.format(projectName, caseFile))
        return 'alreadyExist', caseFile


if __name__ == '__main__':
    excel = create_project('test')
    print(excel)
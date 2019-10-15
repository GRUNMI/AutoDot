# coding:utf-8
__author__ = 'GRUNMI'

from common.CreateProject import create_project
from common.LogOutput import Log
import openpyxl
import xlrd
import time


def case_data(projectName,action,id=1,run='YES',sort=None,sql=None,moduleName='模块名称',caseName='用例名称',
              caseDescription='用例描述',protocol='http://',host=None,method='GET',headers=None,path=None,
              data=None,file=None,replaceID=None,expect=None,checkPoint=None,response=None,transmitID=None,
              transmitTargetID=None,remark=None,createDate=None,updateTime=None,author='GRUNMI',editor='GRUNMI',runResult=None):


    '''

    :param projectName: 项目名称
    :param action: 执行动作，1增 2删 3修改用例数据 4修改用例执行状态 5查
    :param id: 用例id
    :param run: 用例执行
    :param sort: 用例排序
    :param sql: 数据库语法
    :param moduleName: 模块名称
    :param caseName: 用例名称
    :param caseDescription: 用例描述
    :param protocol: 协议
    :param host: 域名
    :param method: 请求方法
    :param headers: 请求数据格式
    :param path: 请求路径
    :param data: 请求数据
    :param file: 请求文件
    :param replaceID: 替换请求动态数据
    :param expect: 预期结果
    :param checkPoint: 校验数据
    :param response: 实际结果
    :param transmitID: 传递的动态id
    :param transmitTargetID: 传递到目标动态id
    :param remark: 备注
    :param createDate: 创建时间
    :param updateTime: 更新时间
    :param author: 作者
    :param editor: 修改人
    :param runResult: 运行状态
    :return:
    '''


    mylogger = Log().get_logger()
    mylogger.info('当前运行文件：{}'.format(__file__))
    mylogger.info('调用运行create_project方法')
    caseFile = create_project(projectName)[1]
    # print(caseFile)
    mylogger.info('打开{}文件'.format(caseFile))

    action = int(action)
    id = int(id)
    # openpyxl操作excel
    wb = openpyxl.load_workbook(caseFile, data_only=True)

    # print(wb.get_sheet_names())
    # print(wb.sheetnames)
    # wb.get_sheet_names() 已弃用，需要使用sheetName
    # xlrd读取excel
    caseFiles = xlrd.open_workbook(caseFile)
    caseSheet = caseFiles.sheet_by_index(0)

    try:
        for ws in wb.sheetnames:
            # count：判断是否包含projectName
            if ws.count(projectName):
                mylogger.info('已匹配到{}的用例表为：{}'.format(projectName, ws))
                # wb.get_sheet_by_name() 已弃用，需要使用wb[]
                ws = wb[ws]
                rows = ws.max_row
                # print(type(rows), '------------------')
                mylogger.info('获取当前{}表的最大行数：{}'.format(ws, rows))
                # 新增
                if action == 1:
                    write_row = str(rows+1)
                    mylogger.info('准备将数据新增在{}表的第{}行'.format(ws, write_row))

                    # 获取最后一行的id值+1
                    if rows > 1:
                        ws['A' + write_row] = int(ws['A' + str(rows)].value) + 1
                    else:
                        ws['A' + write_row] = 1

                    ws['B' + write_row] = run
                    # 减去用例标题一行
                    ws['C' + write_row] = int(write_row) - 1
                    ws['D' + write_row] = sql
                    ws['E' + write_row] = moduleName
                    ws['F' + write_row] = caseName
                    ws['G' + write_row] = caseDescription
                    ws['H' + write_row] = protocol
                    ws['I' + write_row] = host
                    ws['J' + write_row] = method
                    ws['K' + write_row] = headers
                    ws['L' + write_row] = path
                    ws['M' + write_row] = data
                    ws['N' + write_row] = file
                    ws['O' + write_row] = replaceID
                    ws['P' + write_row] = expect
                    ws['Q' + write_row] = checkPoint
                    # ws['R' + write_row] = response
                    ws['S' + write_row] = transmitID
                    ws['T' + write_row] = transmitTargetID
                    # ws['U' + write_row] = remark
                    ws['V' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                    ws['W' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                    ws['X' + write_row] = author
                    ws['Y' + write_row] = editor
                    # ws['Z' + write_row] = runResult
                    mylogger.info('新增在{}表中第{}行的数据成功'.format(ws, write_row))
                    mylogger.info('总共有{}条用例数据'.format(int(write_row)-1))
                # 删除
                elif action == 2:
                    mylogger.info('准备删除{}表中，id为{}的用例数据'.format(ws, id))
                    if rows > 1:
                        try:
                            for id_col, id_title in enumerate(caseSheet.row_values(0)):
                                # 用例的标题
                                if id_title == 'ID':
                                    print(caseSheet.col_values(id_col))
                                    for id_row, id_val in enumerate(caseSheet.col_values(id_col)[1::]):
                                        if int(id_val) == id:

                                            # 修改sort数据
                                            # mylogger.info('准备更新{}表中，sort数据'.format(ws))
                                            for sort_col, sort_title in enumerate(caseSheet.row_values(0)):
                                                if sort_title == 'Sort':

                                                    # 需要删除的id的sort值
                                                    id_sort_value = ws.cell(row=id_row + 2, column=sort_col + 1).value

                                                    for sort_row, sort_val in enumerate(caseSheet.col_values(sort_col)[1::]):
                                                        if int(sort_val) > int(id_sort_value):
                                                            val = ws.cell(row=int(sort_row)+2, column=int(sort_col)+1).value
                                                            mylogger.info(
                                                                '{}表中，修改{}大于{}的sort数据成功'.format(ws, val, id_sort_value))
                                                            val -= 1
                                                            ws.cell(row=int(sort_row) + 2, column=int(sort_col) + 1, value=val)
                                                    break

                                            # 删除数据
                                            mylogger.info('删除{}表中，id为{}的用例数据成功'.format(ws, id))
                                            ws.delete_rows(id_row + 2)
                                            mylogger.info('删除{}表中，第{}条用例数据'.format(ws, id_row + 1))
                                            break

                                    # mylogger.info('{}表中，删除用例数据不成功，id为{}的用例数据不存在'.format(ws, id))
                                    break

                        except Exception as e:
                            mylogger.info('删除{}表中,id为{}的用例报错：{}'.format(ws, id, e))
                    elif rows == 1:
                        mylogger.info('{}表中不存在用例数据，删除用例数据不成功'.format(ws))
                    mylogger.info('总共有{}条用例数据'.format(int(rows) - 1))

                # 修改用例数据
                elif action == 3:
                    mylogger.info('准备修改{}表中，id为{}的用例数据'.format(ws, id))
                    if rows > 1:
                        for id_col, id_title in enumerate(caseSheet.row_values(0)):
                            if id_title == 'ID':
                                for id_row, id_val in enumerate(caseSheet.col_values(id_col)[1::]):
                                    if int(id_val) == id:
                                        write_row = str(id_row+2)
                                        # 获取最后一行的id值+1
                                        # ws['A' + write_row] = int(ws['A' + write_row].value) + 1

                                        # ws['B' + write_row] = run
                                        # 减去用例标题一行
                                        # ws['C' + write_row] = int(write_row) - 1
                                        ws['D' + write_row] = sql
                                        ws['E' + write_row] = moduleName
                                        ws['F' + write_row] = caseName
                                        ws['G' + write_row] = caseDescription
                                        ws['H' + write_row] = protocol
                                        ws['I' + write_row] = host
                                        ws['J' + write_row] = method
                                        ws['K' + write_row] = headers
                                        ws['L' + write_row] = path
                                        ws['M' + write_row] = data
                                        ws['N' + write_row] = file
                                        ws['O' + write_row] = replaceID
                                        ws['P' + write_row] = expect
                                        ws['Q' + write_row] = checkPoint
                                        # ws['R' + write_row] = response
                                        ws['S' + write_row] = transmitID
                                        ws['T' + write_row] = transmitTargetID
                                        # ws['U' + write_row] = remark
                                        # ws['V' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                                        ws['W' + write_row] = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                                        ws['X' + write_row] = author
                                        ws['Y' + write_row] = editor
                                        # ws['Z' + write_row] = runResult
                                        mylogger.info('修改{}表中，id为{}的用例数据成功'.format(ws, id))
                                        break
                        #         mylogger.info('修改{}表中，id为{}的用例数据不存在，修改失败'.format(ws, id))
                        # mylogger.info('修改{}表中，不存在标题为ID的'.format(ws))

                    else:
                        mylogger.info('{}表中不存在用例数据，删除用例数据不成功'.format(ws))

                # 修改是否执行用例
                elif action == 4:
                    mylogger.info('准备修改{}表中，id为{}的用例执行状态为：{}'.format(ws, id, run))
                    if rows > 1:
                        for id_col, id_title in enumerate(caseSheet.row_values(0)):
                            if id_title == 'ID':
                                for id_row, id_val in enumerate(caseSheet.col_values(id_col)[1::]):
                                    if int(id_val) == id:
                                        write_row = str(id_row+2)
                                        if ws['B' + write_row].value == run:
                                            mylogger.info('修改{}表中，id为{}的用例执行状态数据未做改变，不需要更新'.format(ws, id))
                                            break
                                        else:
                                            ws['B' + write_row] = run
                                            mylogger.info('修改{}表中，id为{}的用例执行状态数据更新成功'.format(ws, id))
                                            break
                                # mylogger.info('修改{}表中，id为{}的用例数据不存在，更新失败'.format(ws, id))
                                break
                        # mylogger.info('修改{}表中，不存在标题为ID的'.format(ws))
                    else:
                        mylogger.info('{}表中不存在用例数据，删除用例数据不成功'.format(ws))

                # 查询
                elif action == 5:
                    mylogger.info('准备查询{}表中，id为{}的用例数据'.format(ws, id))
                    if rows > 1:
                        for id_col, id_title in enumerate(caseSheet.row_values(0)):
                            if id_title == 'ID':
                                for id_row, id_val in enumerate(caseSheet.col_values(id_col)[1::]):
                                    if int(id_val) == id:
                                        find_row = str(id_row+2)

                                        ID=ws['A' + find_row].value
                                        Run=ws['B' + find_row].value
                                        Sort=ws['C' + find_row].value
                                        Sql=ws['D' + find_row].value
                                        ModuleName=ws['E' + find_row].value
                                        CaseName=ws['F' + find_row].value
                                        CaseDescription=ws['G' + find_row].value
                                        Protocol=ws['H' + find_row].value
                                        Host=ws['I' + find_row].value
                                        Method=ws['J' + find_row].value
                                        Headers=ws['K' + find_row].value
                                        Path=ws['L' + find_row].value
                                        Data=ws['M' + find_row].value
                                        Files=ws['N' + find_row].value
                                        ReplaceID=ws['O' + find_row].value
                                        Expect=ws['P' + find_row].value
                                        CheckPoint=ws['Q' + find_row].value
                                        Response=ws['R' + find_row].value
                                        TransmitID=ws['S' + find_row].value
                                        TransmitTargetID=ws['T' + find_row].value
                                        Remark=ws['U' + find_row].value
                                        CreateDate=ws['V' + find_row].value
                                        UpdateTime=ws['W' + find_row].value
                                        Author=ws['X' + find_row].value
                                        Editor=ws['Y' + find_row].value
                                        RunResult=ws['Z' + find_row].value


                                        find_result = {
                                            'projectname': projectName,
                                            'id': ID,
                                            'run': Run,
                                            'sort': Sort,
                                            'sql': Sql,
                                            'moduleName': ModuleName,
                                            'caseName': CaseName,
                                            'caseDescription': CaseDescription,
                                            'protocol': Protocol,
                                            'host': Host,
                                            'method': Method,
                                            'headers': Headers,
                                            'path': Path,
                                            'data': Data,
                                            'files': Files,
                                            'replaceID': ReplaceID,
                                            'expect': Expect,
                                            'checkPoint': CheckPoint,
                                            'result': Response,
                                            'transmitID': TransmitID,
                                            'transmitTargetID': TransmitTargetID,
                                            'remark': Remark,
                                            'createDate': CreateDate,
                                            'updateTime': UpdateTime,
                                            'author': Author,
                                            'editor': Editor,
                                            'runResult': RunResult,
                                        }
                                        mylogger.info('查询{}表中，id为{}的用例数据成功：{}'.format(ws, id, find_result))
                                        wb.close()
                                        mylogger.info('关闭{}工作簿成功'.format(projectName))
                                        return find_result
                                mylogger.info('查询{}表中，id为{}的用例数据不存在'.format(ws, id))
                                return 'noFind'

                    else:
                        mylogger.info('{}表中，不存在用例数据'.format(ws))

                mylogger.info('准备保存{}表'.format(ws))
                wb.save(caseFile)
                mylogger.info('保存{}表成功，准备关闭{}工作簿'.format(ws, projectName))
                break

    except Exception as e:
        mylogger.info('不可预知的错误：{}'.format(e))
    wb.close()
    mylogger.info('关闭{}工作簿成功'.format(projectName))

if __name__ == '__main__':
    # case_data(projectName='test', action=1)
    # case_data(projectName='test', action=2, id=4)
    # case_data(projectName='test', action=3,id=5,sql=111,moduleName='模块名称',caseName='用例名称',
    #           caseDescription='用例描述',protocol='http://',host=None,method='GET',headers=None,path=None,
    #           data=None,file=None,replaceID=None,expect=None,checkPoint=None,transmitID=None,
    #           transmitTargetID=None,author='GRUNMI',editor='GRUNMI')
    # case_data(projectName='test', action=4, id=6, run='NO')
    case_data(projectName='test',action=5,id=2)